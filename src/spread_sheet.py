from openpyxl import load_workbook
from openpyxl.comments import Comment
import re

_wb = None

_sheetMap = {
    'name': 'B6', 'soo': 'B7', 'mat_no': 'D6',
    'sex': 'G7', 'marital': 'E7', 'session': 'C9',
    'hod': 'B22', 'dept': 'A3', 'faculty': 'A2'
}

class _LeadData(object):
    hod = "22"
    def __init__(self):
        super().__init__()

class _Level(object):
    def __init__(self, level, session, ws, lead_data,  is_lead=False, is_tail=False):
        super().__init__()
        self.results = [[],[]]
        self.total_shift = 0
        self.ws = ws
        self.level = level
        self.session = session
        self.lead_data = lead_data
        self.is_lead = is_lead
        self.is_tail = is_tail
        self.tables = [ws.tables.get('S' + str(level) + '.1'), ws.tables.get('S' + str(level) + '.2')]
    
    def add_result(self, result, semester):
        self.results[semester - 1].append(result)

    def commit(self):
        ws = self.ws
        ws[_sheetMap['session']] = str(self.session - 1) + '/' + str(self.session)
        
        row_shift1 = len(self.results[0]) - 1
        ref1 = self.tables[0].ref
        if row_shift1 > 0:
            ws.insert_rows(int(self._split_ref(ref1)[4]), row_shift1)
            self.tables[0].ref = self._shift_range(ref1, row_shift1)
        else:
            row_shift1 = 0

        row_shift2 = len(self.results[1]) - 1
        ref2 = self.tables[1].ref
        if row_shift1 > 0 or row_shift2 > 0:
            if row_shift2 > 0:
                ws.insert_rows(int(self._split_ref(ref2)[4]) + row_shift1, row_shift2)
            else:
                row_shift2 = 0
            self.tables[1].ref = self._shift_range(ref2, row_shift1 + row_shift2, row_shift1)
        
        refs = [
            int(self._split_ref(self.tables[0].ref)[2]) + 1,
            int(self._split_ref(self.tables[1].ref)[2]) + 1
        ]
        for c in range(2):
            i = 0
            for result in self.results[c]:
                top = refs[result['sem'] - 1]
                ws['A' + str(i + top)] = result['code']
                ws['B' + str(i + top)] = result['title']
                ws['C' + str(i + top)] = result['cu']
                ws['D' + str(i + top)] = result.get('score')
                for c in range(3, 8):
                    ws.cell(i + top, c).style = ws.cell(top, c).style
                for c in range(5, 8):
                    ws.cell(i + top, c).value = ws.cell(top, c).value
                if result.get('comment') != '' and result.get('comment') != None:
                    ws.cell(i + top, 4).comment = Comment('Revisions:\n' + result.get('comment'), 'Auto', width=300)
                i += 1
        total_shift = row_shift1 + row_shift2
        self.total_shift = total_shift
        if self.is_lead:
            self.lead_data.hod = str(22 + total_shift)
        if self.is_tail:
            ws['G' + str(23 + total_shift)] = ws['G' + str(23 + total_shift)].value.replace('20', str(20 + total_shift))

    def _shift_range(self, range, bottom, top=0):
        rng = self._split_ref(range)
        return rng[1] + str(int(rng[2]) + top) + ':' + rng[3] + str(int(rng[4]) + bottom)

    def _split_ref(self, ref):
        return re.split('([A-Z]+)(\d+):([A-Z]+)(\d+)', ref)

    def finish(self):
        if not self.is_lead:
            hod_cell = str(22 + self.total_shift)
            self.ws['B' + hod_cell] = self.ws['B' + hod_cell].value.replace('22', self.lead_data.hod)

def generate_spread_sheet(user, results, courses, filename):
    global _wb
    _wb = load_workbook('static/excel/spreadsheet_template.xlsx')

    user['name'] = (user['last_name'].upper() + ', ' + 
        user['first_name'].capitalize() + ' ' + user['other_name'].capitalize())
    for key in user.keys():
        if _sheetMap.get(key) != None:
            _wb['L100'][_sheetMap[key]] = user[key]
    
    level_status = { 'sessions': [0], 'last_sem': 101 }
    result_map = {}

    # step 1: remove carry-overs
    for result in results:
        result.update(courses[result['courseCode']])
        result.update({'_session': result['session'], 'comment': ''})
        map = result_map.get(result['courseCode'])
        if  level_status['sessions'][-1] != result['session']:
            level_status['sessions'].append(result['session'])
        if result['level'] + result['sem'] > level_status['last_sem']:
            level_status['last_sem'] = result['level'] + result['sem']
        if result['score'] < 40:
            result['cu'] = 0
        if map == None or (map['score'] < 40 and result['session'] > map['session']):
            if map != None:
                result['comment'] = (map['comment'] + '[ session: ' + str(map['session'] - 1) + '/' 
                    + str(map['session']) + ', score: ' + str(map['score']) + ']\n')
                result['_session'] = result['session'] + 0.4
            result_map[result['courseCode']] = result
        else:
            result_map[result['courseCode']]['comment'] = (map['comment'] + 'flag* [ session: ' + str(result['session'] - 1) + '/' 
                + str(result['session']) + ', score: ' + str(result['score']) + ']\n')

    # step 2: add missing courses till last semester    
    for key in courses.keys():
        if result_map.get(key) == None and courses[key]['level'] + courses[key]['sem'] <= level_status['last_sem']:
            result_map[key] = courses[key]
            session = level_status['sessions'][int(courses[key]['level']/100)]
            result_map[key].update({'courseCode': key, 'cu': None, '_session': session + 0.5, 'session': session })

    final_results = []
    final_results.extend(result_map.values())
    final_results.sort(key = lambda i: (i['_session'], i['code']))

    # step 3: write reuluts to sheet
    _write_results(final_results, level_status)
    
    # step 4: remove unused sheets
    for sheet in _wb.worksheets:
        if sheet[_sheetMap['session']].value == None:
            _wb.remove(sheet)
    _wb.save(filename)
    _wb.close()
    _wb = None


def _write_results(results, status):
    levels = {}
    lead_data = _LeadData()
    for result in results:
        level = status['sessions'].index(result['session'])
        sem = result['sem']
        if levels.get(str(level)) == None:
            levels[str(level)] = _Level(level, result['session'], _wb['L' + str(level * 100)], lead_data, is_lead=level == 1, is_tail=level>=5)
        levels[str(level)].add_result(result, sem)
    for level in levels.values():
        level.commit()
    for level in levels.values():
        level.finish()

if __name__ == '__main__':
    # Run a test using sample data
    from sample_data.result import user, result
    from sample_data.courses import MEG, MCT
    
    # sort results by only session
    result.sort(key = lambda i: (i['session']))
    courses = MEG

    if user['department'] == 'MCT':
        courses = MCT
    generate_spread_sheet(user, result, courses, filename='output/sample_spreadsheet.xlsx')

