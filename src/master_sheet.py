from openpyxl import load_workbook
import uuid, re

class MasterSheet(object):
    def __init__(self, filename, courseCode='', session=''):
        super().__init__()
        self.batchId = self.getBatchId()
        self.courseId = self.rnd_id()
        self.filename = filename
        self.anchor = None
        self.topLeft = [0, 0]
        self.bottomRight = [0, 0]
        self.headerMap = []
        self.data_rows = []
        self.session = session
        self.courseCode = courseCode

    def is_Anchor(self, value):
        return re.search('matric', str(value).lower()) != None

    def map_Header(self, value):
        if re.search('matric', str(value).lower()) != None:
            return 'mat_no'
        elif re.search('(total|score)', str(value).lower()) != None:
            return 'score'
        return 'annotation'

    def getBatchId(self):
        return self.rnd_id()

    def rnd_id(self):
        return uuid.uuid4().hex

    def getResultId(self, mat_no):
        return str(self.session) + '-' + self.courseId + '-' + mat_no.replace('/', '-')

    def _has_rows(self):
        return self.bottomRight[0] > self.topLeft[0]

    def _parse_header(self):
        global anchor, session, courseCode
        for r in range(1, 21):
            for c in range(1, 21):
                cell = self.ws.cell(r, c).value
                if self.is_Anchor(cell):
                    self.anchor = [r, c]
                    break
                elif re.search('session', str(cell).lower()) != None:
                    s = self._peek_right('^(_){0,1}((\d){2}|(\d){4})/((\d){2}|(\d){4})(_){0,1}$', r, c)
                    y2 = s.strip().replace('_','').split('/')[1]
                    if len(y2) == 2:
                        y2 = '20' + y2
                    self.session = int(y2)
                elif re.search('course (code|no)', str(cell).lower()) != None:
                    code = self._peek_right('^(_){0,1}[A-z]{3}(\s|_){0,1}\d{3}\.\d(_){0,1}$', r, c)
                    c1 = code.replace('_', '').replace(' ', '').lower()
                    c2 = re.split('([a-z]{3})(\d{3})\.(\d)', c1)
                    self.courseCode = c2[1] + '_' + c2[2] + '_' + c2[3]
                # TODO parse other headers
            if self.anchor != None:
                break

    def _peek_right(self, pattern, row, column):
        for c in range(column + 1, column + 4):
            cell = self.ws.cell(row, c).value
            if re.search(pattern, str(cell).strip()) != None:
                return cell

    def _parse_sheet(self):
        self._parse_header()
        if self.anchor != None:
            self._go_left()
            self._go_right()
            self._go_down()
            self._map_Headers()
            if self._has_rows():
                self._get_rows()

    def _get_rows(self):
        for ur in range(self.topLeft[0] + 1, self.bottomRight[0] + 1):
            row = {
                'batchId': self.batchId, 'session': self.session,
                'courseId': self.courseId, 'courseCode': self.courseCode,
                'mat_no': '', 'score': ''
                }
            annotation = []
            for uc in range(self.topLeft[1], self.bottomRight[1] + 1):
                if self.headerMap[uc - self.topLeft[1]] != 'annotation':
                    row[self.headerMap[uc - self.topLeft[1]]] = self.sanitize(self.ws.cell(ur, uc).value)
                else:
                    annotation.append({
                        'key': self.sanitize(self.ws.cell(self.anchor[0], uc).value),
                        'value': self.sanitize(self.ws.cell(ur, uc).value)
                    })
            row['annotation'] = str(annotation)
            row['mat_no'] = row['mat_no'].upper()
            # TODO tests and sanitize (mat number, score), yada yada yada!

            row['resultId'] = self.getResultId(row['mat_no'])
            self.data_rows.append(row)

    def sanitize(self, value):
        if value == None:
            return ''
        elif type(value) == str:
            return value.strip()
        else:
            return value

    def _map_Headers(self):
        for h in range(self.topLeft[1], self.bottomRight[1] + 1):
            self.headerMap.append(self.map_Header(self.ws.cell(self.anchor[0], h).value))
            
    def _go_left(self):
        self.topLeft[1] = self.anchor[1]
        for l in range(self.anchor[1] -1, 0, -1):
            if self.ws.cell(self.anchor[0], l).value == None:
                break
            else:
                self.topLeft[1] = l
    
    def _go_right(self):
        self.bottomRight[1] = self.anchor[1]
        while True:
            self.bottomRight[1] = self.bottomRight[1] + 1
            if self.ws.cell(self.anchor[0], self.bottomRight[1]).value == None:
                self.bottomRight[1] = self.bottomRight[1] - 1
                break
    
    def _go_down(self):
        self.topLeft[0] = self.anchor[0]
        self.bottomRight[0] = self.anchor[0]
        while True:
            self.bottomRight[0] = self.bottomRight[0] + 1
            if self.ws.cell(self.bottomRight[0], self.anchor[1]).value == None:
                self.bottomRight[0] = self.bottomRight[0] - 1
                break

    def get_results(self):
        self.wb = load_workbook(self.filename, data_only=True)
        for sheet in self.wb.worksheets:
            self.ws = sheet
            self.anchor = None
            self._parse_sheet()
        self.wb.close()
        self.wb = None
        return self.data_rows

if __name__ == '__main__':
    # Run a test using sample master sheet
    master = MasterSheet('static/excel/ENG301.1.xlsx', courseCode='chm_130_1', session=2019)
    results = master.get_results()
    batch = master.batchId
    code = master.courseCode
    session = master.session

    for data in results:
        print(data)
